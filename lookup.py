from openpyxl import load_workbook
import re

dictionary = list()

for row in load_workbook("dictionary.xlsx", read_only=True)["overall"].iter_rows(min_row=4, min_col=3, max_col=3):
    for cell in row:
        dictionary.extend(set(cell.value.lower().split('/')))

dictionary = sorted(dictionary)

open("dictionary.txt", 'w').write('\n'.join(word for word in dictionary))